from flask import Flask, request, send_file
from flask_cors import CORS
import openpyxl
import shutil, io, os, tempfile
from datetime import datetime

app = Flask(__name__)
CORS(app)

BASE     = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE, 'template.xlsx')


def make_excel(data):
    tax_mode = data.get('tax_mode', 'with')
    items    = data.get('items', [])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    shutil.copy2(TEMPLATE, tmp.name)

    wb = openpyxl.load_workbook(tmp.name)
    ws = wb['報價單']

    ws['C8'].value  = data.get('date', '')
    ws['C9'].value  = data.get('client', '')
    ws['C10'].value = data.get('contact', '')

    # 清空商品行
    for r in range(16, 36):
        ws[f'B{r}'].value = None
        ws[f'D{r}'].value = None
        ws[f'E{r}'].value = None
        ws[f'F{r}'].value = None  # 不用公式，直接清空

    # 填入商品，F 欄直接寫數字
    subtotal = 0
    for i, item in enumerate(items[:19]):
        r     = 16 + i
        price = float(item.get('price') or 0)
        qty   = float(item.get('qty')   or 0)
        sub   = price * qty
        subtotal += sub

        ws[f'B{r}'].value = item.get('desc', '')
        ws[f'D{r}'].value = price
        ws[f'E{r}'].value = qty
        ws[f'F{r}'].value = sub   # ✅ 直接寫數字

    # 以下空白
    blank_row = 16 + len(items)
    if blank_row <= 35:
        ws[f'B{blank_row}'].value = '以下空白'

    # 備註
    note = data.get('note', '')
    ws['B37'].value = f'備註:{note}' if note else '備註:'

    # 小計 / 稅金 / 總計 直接寫數字
    tax   = round(subtotal * 0.05) if tax_mode == 'with' else 0
    total = subtotal + tax

    ws['E37'].value = '小計'
    ws['F37'].value = subtotal   # ✅ 直接寫數字

    if tax_mode == 'without':
        ws['E38'].value = None
        ws['F38'].value = None
    else:
        ws['E38'].value = '稅金'
        ws['F38'].value = tax    # ✅ 直接寫數字

    ws['E39'].value = '總計金額'
    ws['F39'].value = total      # ✅ 直接寫數字

    wb.save(tmp.name)
    return tmp.name


@app.route('/')
def index():
    return send_file(os.path.join(BASE, 'index.html'))

@app.route('/ping')
def ping():
    return 'ok'


@app.route('/generate', methods=['POST'])
def generate():
    data      = request.json
    xlsx_path = make_excel(data)

    with open(xlsx_path, 'rb') as f:
        content = f.read()
    os.unlink(xlsx_path)

    client    = data.get('client', '未命名')
    today     = datetime.now().strftime('%Y%m%d')
    tax_label = '含稅' if data.get('tax_mode','with') == 'with' else '未稅'
    filename  = f'正陽資訊報價單_{client}_{tax_label}_{today}.xlsx'

    return send_file(
        io.BytesIO(content),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/preview_pdf', methods=['POST'])
def preview_pdf():
    data     = request.json
    tax_mode = data.get('tax_mode', 'with')
    items    = data.get('items', [])
    note     = data.get('note', '')
    client   = data.get('client', '')
    date     = data.get('date', '')
    contact  = data.get('contact', '')

    subtotal = sum(float(i.get('price') or 0) * float(i.get('qty') or 0) for i in items)
    tax      = round(subtotal * 0.05) if tax_mode == 'with' else 0
    total    = subtotal + tax

    item_rows = ''
    for item in items:
        price = float(item.get('price') or 0)
        qty   = float(item.get('qty') or 0)
        sub   = price * qty
        item_rows += f'''<tr>
            <td class="desc">{item.get("desc","")}</td>
            <td class="num">${price:,.0f}</td>
            <td class="center">{qty:g}</td>
            <td class="num">${sub:,.0f}</td>
        </tr>'''

    tax_row = f'<tr><td class="sum-label">稅金</td><td class="sum-val">${tax:,.0f}</td></tr>' if tax_mode == 'with' else ''

    html = f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>正陽資訊報價單 - {client}</title>
<style>
  @page {{ size: A4; margin: 18mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "標楷體","DFKai-SB","BiauKai",serif; font-size: 12pt; color: #000; }}
  .title {{ font-size: 22pt; font-weight: bold; text-align: center; padding: 4pt 0 2pt; }}
  .addr-row {{ display: flex; justify-content: space-between; align-items: baseline; padding: 2pt 0 4pt; border-bottom: 1.5pt solid #000; }}
  .addr {{ font-size: 10pt; }}
  .unified {{ font-size: 14pt; font-weight: bold; }}
  .info-table {{ width: 100%; margin: 6pt 0; border-collapse: collapse; font-size: 11pt; }}
  .info-table td {{ padding: 3pt 4pt; }}
  .info-table .label {{ width: 22%; white-space: nowrap; }}
  .info-table .val {{ border-bottom: 0.5pt solid #aaa; width: 28%; }}
  .items-table {{ width: 100%; border-collapse: collapse; margin: 6pt 0; font-size: 11pt; }}
  .items-table thead tr {{ background: #2ED06E; color: #fff; }}
  .items-table th {{ padding: 5pt 6pt; text-align: center; font-weight: bold; }}
  .items-table td {{ padding: 4pt 6pt; border: 0.4pt solid #ccc; }}
  .items-table .desc {{ text-align: left; }}
  .items-table .num  {{ text-align: right; }}
  .items-table .center {{ text-align: center; }}
  .items-table tbody tr:nth-child(even) {{ background: #f5f5f5; }}
  .bottom-row {{ display: flex; justify-content: space-between; align-items: flex-start; margin-top: 4pt; }}
  .note-box {{ width: 52%; font-size: 10pt; color: #5D7079; padding-top: 3pt; }}
  .sum-box {{ width: 46%; }}
  .sum-box table {{ width: 100%; border-collapse: collapse; font-size: 11pt; color: #5D7079; }}
  .sum-box .sum-label {{ text-align: right; padding: 3pt 4pt; font-weight: bold; }}
  .sum-box .sum-val   {{ text-align: right; padding: 3pt 4pt; }}
  .sum-box .total-row td {{ font-size: 14pt; font-weight: bold; border-top: 1.5pt solid #5D7079; padding-top: 5pt; }}
  .terms {{ margin-top: 10pt; border-top: 0.5pt solid #ccc; padding-top: 6pt; font-size: 8pt; color: #5D7079; line-height: 1.6; }}
  .terms-title {{ font-size: 9pt; color: #333; font-weight: bold; margin-bottom: 3pt; }}
  .print-btn {{ position: fixed; bottom: 20px; right: 20px; background: #1a6b3a; color: #fff; border: none; border-radius: 50px; padding: 14px 28px; font-size: 16px; cursor: pointer; box-shadow: 0 4px 14px rgba(0,0,0,0.25); font-family: inherit; }}
  @media print {{ .print-btn {{ display: none; }} }}
</style>
</head>
<body>
<div class="title">正陽資訊有限公司報 價 單</div>
<div class="addr-row">
  <span class="addr"> 淡水門市部:新北市淡水區中山北路一段119巷3號</span>
  <span class="unified">統編:54840749</span>
</div>
<table class="info-table">
  <tr>
    <td class="label">報價日期 ：</td><td class="val">{date}</td>
    <td style="width:4%"></td>
    <td class="label">專責人員 :</td><td class="val" style="text-align:center">曹正岳 0983-822-631</td>
  </tr>
  <tr>
    <td class="label">客戶名稱 ：</td><td class="val">{client}</td>
    <td></td>
    <td class="label">聯 絡 人 :</td><td class="val">{contact}</td>
  </tr>
</table>
<table class="items-table">
  <thead>
    <tr>
      <th style="width:42%;text-align:left;padding-left:8pt">描述</th>
      <th style="width:18%">單價</th>
      <th style="width:13%">數量</th>
      <th style="width:27%">項目合計</th>
    </tr>
  </thead>
  <tbody>{item_rows}</tbody>
</table>
<div class="bottom-row">
  <div class="note-box">備註:{note}</div>
  <div class="sum-box">
    <table>
      <tr><td class="sum-label">小計</td><td class="sum-val">${subtotal:,.0f}</td></tr>
      {tax_row}
      <tr class="total-row"><td class="sum-label">總計金額</td><td class="sum-val">${total:,.0f}</td></tr>
    </table>
  </div>
</div>
<div class="terms">
  <div class="terms-title">條款及細則</div>
  1. 報價單自報價日起 30 日內有效，請確認報價單之內容與交期，簽章回傳後，視為正式訂購單。<br>
  2. 報價單之產品型號及規格經正式訂購後即無法取消或退換貨處理。<br>
  3. 買方如對本報價單有所修改，未經本公司簽名回傳確認，本報價單即歸於無效。<br>
  4. 本公司就本採購案，非屬公職人員利益衝突迴避法第2條及第3條所稱公職人員或關係人。
</div>
<button class="print-btn" onclick="window.print()">列印 / 儲存 PDF</button>
</body>
</html>'''

    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
